import tkinter as tk
from tkinter import messagebox, filedialog, simpledialog
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import os
import sys
import webbrowser
from pathlib import Path
import pandas as pd
import numpy as np
import threading
import openpyxl
from openpyxl.utils import get_column_letter

import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk

import analytics_engine as analyzer
import fp_manager
import trend_engine
import correlation_engine
import cmdb_exporter
import network_visualizer
import config_loader
import gvm_connector
import executive_reporter

SCRIPT_DIR = Path(__file__).parent.absolute()
REPORTS_DIR = SCRIPT_DIR / "reports"
REPORTS_DIR.mkdir(exist_ok=True)

class OpenVASApp:
    def __init__(self, root):
        self.root = root
        self.root.title("🛡️ OpenVAS Pro Suite — Modern Risk-Based Vulnerability Management (v4.0)")
        self.root.geometry("1420x900")
        self.config = config_loader.get_config()

        # Track sorting states per treeview: {col_name: reverse_boolean}
        self.tree_sort_states = {}

        # Top Header Bar
        self.top_bar = ttk.Frame(root, bootstyle="dark", padding=(15, 10))
        self.top_bar.pack(fill=tk.X, side=tk.TOP)

        brand_lbl = ttk.Label(self.top_bar, text="🛡️ OPENVAS PRO SUITE", font=("Segoe UI", 12, "bold"), bootstyle="inverse-dark")
        brand_lbl.pack(side=tk.LEFT)

        sub_lbl = ttk.Label(self.top_bar, text=" | Risk-Based Vulnerability Management Platform (v4.0)", font=("Segoe UI", 9), bootstyle="inverse-dark")
        sub_lbl.pack(side=tk.LEFT, padx=5)

        self.status_pill = ttk.Label(self.top_bar, text="🟢 Status: Ready", font=("Segoe UI", 9, "bold"), bootstyle="inverse-dark")
        self.status_pill.pack(side=tk.RIGHT, padx=10)

        # Main Workspace Split (Sidebar Left + Content Canvas Right)
        self.workspace = ttk.Frame(root)
        self.workspace.pack(fill=tk.BOTH, expand=True)

        self.sidebar = ttk.Frame(self.workspace, width=220, padding=10, bootstyle="secondary")
        self.sidebar.pack(side=tk.LEFT, fill=tk.Y)

        self.content_area = ttk.Frame(self.workspace, padding=12)
        self.content_area.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        self.nav_buttons = {}
        self.views = {}

        self.views["reports"] = ttk.Frame(self.content_area)
        self.views["analytics"] = ttk.Frame(self.content_area)
        self.views["fp"] = ttk.Frame(self.content_area)
        self.views["trend"] = ttk.Frame(self.content_area)
        self.views["correlation"] = ttk.Frame(self.content_area)
        self.views["assets"] = ttk.Frame(self.content_area)
        self.views["sla"] = ttk.Frame(self.content_area)

        self.tab_reports = self.views["reports"]
        self.tab_analytics = self.views["analytics"]
        self.tab_fp = self.views["fp"]
        self.tab_trend = self.views["trend"]
        self.tab_correlation = self.views["correlation"]
        self.tab_assets = self.views["assets"]
        self.tab_sla = self.views["sla"]

        nav_structure = [
            ("HEADER", "SCAN MANAGEMENT"),
            ("reports", "📁 Reports & GVM Live"),
            ("HEADER", "THREAT & RISK"),
            ("analytics", "📊 Analytics & Threat Intel"),
            ("fp", "🏷️ False Positives Audit"),
            ("trend", "📈 Trend Heatmaps"),
            ("correlation", "🌐 Master Targets"),
            ("HEADER", "ASSET COMPLIANCE"),
            ("assets", "🗺️ Topology & CMDB"),
            ("sla", "⏱️ Remediation SLAs"),
        ]

        for item in nav_structure:
            if item[0] == "HEADER":
                lbl = ttk.Label(self.sidebar, text=item[1], font=("Segoe UI", 8, "bold"), bootstyle="inverse-secondary")
                lbl.pack(anchor=tk.W, pady=(14, 4), padx=6)
            else:
                key, text = item[0], item[1]
                btn = ttk.Button(self.sidebar, text=text, bootstyle="outline-light", width=24, command=lambda k=key: self.select_view(k))
                btn.pack(fill=tk.X, pady=3)
                self.nav_buttons[key] = btn

        # Setup Content Layouts
        self.setup_reports_tab()
        self.setup_analytics_tab()
        self.setup_fp_tab()
        self.setup_trend_tab()
        self.setup_correlation_tab()
        self.setup_assets_tab()
        self.setup_sla_tab()

        self.active_view = None
        self.select_view("reports")

        # Bottom Status Bar with Progress Bar
        self.status_frame = ttk.Frame(root, relief=tk.SUNKEN, padding=4)
        self.status_frame.pack(side=tk.BOTTOM, fill=tk.X)

        self.status_var = tk.StringVar(value=f"System Ready. Target Directory: {REPORTS_DIR}")
        self.status_label = ttk.Label(self.status_frame, textvariable=self.status_var, font=("Segoe UI", 9))
        self.status_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.progress_bar = ttk.Progressbar(self.status_frame, orient=tk.HORIZONTAL, mode="indeterminate", length=240)

        self.gvm_reports_data = {}
        self.current_risk_df = pd.DataFrame()
        self.current_master_df = pd.DataFrame()

    def select_view(self, view_key):
        for key, btn in self.nav_buttons.items():
            if key == view_key:
                btn.config(bootstyle="primary")
            else:
                btn.config(bootstyle="outline-light")

        for key, frame in self.views.items():
            if key == view_key:
                frame.pack(fill=tk.BOTH, expand=True)
            else:
                frame.pack_forget()

        self.active_view = view_key

    def log(self, msg):
        self.status_var.set(msg)
        self.root.update_idletasks()

    def start_progress(self, mode="indeterminate"):
        self.progress_bar.config(mode=mode)
        self.progress_bar.pack(side=tk.RIGHT, padx=5)
        if mode == "indeterminate":
            self.progress_bar.start(10)
        else:
            self.progress_bar["value"] = 0
        self.root.update_idletasks()

    def update_progress(self, current, total, msg=None):
        if self.progress_bar["mode"] == "determinate":
            pct = (current / total) * 100 if total > 0 else 0
            self.progress_bar["value"] = pct
        if msg:
            self.log(msg)
        self.root.update_idletasks()

    def stop_progress(self):
        if self.progress_bar["mode"] == "indeterminate":
            self.progress_bar.stop()
        self.progress_bar.pack_forget()
        self.root.update_idletasks()

    # ---------------------------------------------------------
    # EXCEL SORTING & EXPORT UTILITIES
    # ---------------------------------------------------------
    def sort_treeview_column(self, tree, col, df_source_getter, render_fn):
        """Interactive column sorting handler when clicking column header."""
        df = df_source_getter()
        if df.empty or col not in df.columns:
            return

        state_key = (id(tree), col)
        reverse = not self.tree_sort_states.get(state_key, False)
        self.tree_sort_states[state_key] = reverse

        # Sort numeric or string smartly
        try:
            sorted_df = df.sort_values(by=col, ascending=not reverse, key=lambda s: pd.to_numeric(s, errors='ignore'))
        except Exception:
            sorted_df = df.sort_values(by=col, ascending=not reverse)

        render_fn(sorted_df)

    def export_dataframe_to_excel(self, df, default_name="OpenVAS_Export.xlsx"):
        """Exports any DataFrame to an auto-formatted Excel workbook (.xlsx)."""
        if df.empty:
            messagebox.showwarning("Empty Data", "There is no data to export.")
            return

        save_path = filedialog.asksaveasfilename(
            title="Save as Excel Workbook",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Workbooks", "*.xlsx"), ("All Files", "*.*")]
        )
        if not save_path:
            return

        try:
            self.start_progress("indeterminate")
            with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Data Export")
                ws = writer.sheets["Data Export"]
                # Bold headers
                for cell in ws[1]:
                    cell.font = cell.font.copy(bold=True)
                # Auto column width
                for i, col in enumerate(df.columns, start=1):
                    col_letter = get_column_letter(i)
                    max_len = max(df[col].astype(str).map(len).max() if len(df) else 0, len(str(col))) + 3
                    ws.column_dimensions[col_letter].width = min(max_len, 60)
                ws.freeze_panes = "A2"

            messagebox.showinfo("Excel Export Complete", f"Successfully exported to Excel:\n{save_path}")
            self.log(f"Exported data to Excel: {save_path}")
            if messagebox.askyesno("Open File", "Would you like to open the exported Excel file now?"):
                webbrowser.open(f"file:///{save_path}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to save Excel file: {e}")
        finally:
            self.stop_progress()

    # ---------------------------------------------------------
    # TAB 1: REPORTS MANAGEMENT & CONTROL PANEL
    # ---------------------------------------------------------
    def setup_reports_tab(self):
        frame = self.tab_reports

        action_bar = ttk.LabelFrame(frame, text="⚡ Quick Navigation & Actions (Move Selected Scans Between Tabs)", padding=8)
        action_bar.pack(fill=tk.X, padx=10, pady=(10, 5))

        ttk.Button(action_bar, text="📊 Send to Analytics", bootstyle="info", command=self.action_send_to_analytics).pack(side=tk.LEFT, padx=4)
        ttk.Button(action_bar, text="📈 Send to Trend Heatmap", bootstyle="secondary", command=self.action_send_to_trend).pack(side=tk.LEFT, padx=4)
        ttk.Button(action_bar, text="🌐 Send to Master Targets", bootstyle="warning", command=self.action_send_to_correlation).pack(side=tk.LEFT, padx=4)
        ttk.Button(action_bar, text="🗺️ Send to Network Topology", bootstyle="secondary", command=self.action_send_to_topology).pack(side=tk.LEFT, padx=4)
        ttk.Button(action_bar, text="📄 Generate Executive Report", bootstyle="success", command=self.generate_executive_report).pack(side=tk.RIGHT, padx=4)

        paned = ttk.Panedwindow(frame, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        left_frame = ttk.Frame(paned, padding=10)
        paned.add(left_frame, weight=1)

        ttk.Label(left_frame, text="SSH Connection", font=("Segoe UI", 10, "bold")).pack(anchor=tk.W, pady=(0, 2))
        
        ttk.Label(left_frame, text="SSH IP:").pack(anchor=tk.W)
        self.ssh_ip = ttk.Entry(left_frame, width=26)
        self.ssh_ip.insert(0, self.config['ssh']['host'])
        self.ssh_ip.pack(pady=(0, 4))

        ttk.Label(left_frame, text="SSH Username:").pack(anchor=tk.W)
        self.ssh_user = ttk.Entry(left_frame, width=26)
        self.ssh_user.insert(0, self.config['ssh']['username'])
        self.ssh_user.pack(pady=(0, 4))

        ttk.Label(left_frame, text="SSH Password:").pack(anchor=tk.W)
        self.ssh_pass = ttk.Entry(left_frame, width=26, show="*")
        self.ssh_pass.pack(pady=(0, 8))

        ttk.Label(left_frame, text="GVM Authentication", font=("Segoe UI", 10, "bold")).pack(anchor=tk.W, pady=(4, 2))

        ttk.Label(left_frame, text="GVM Username:").pack(anchor=tk.W)
        self.gvm_user = ttk.Entry(left_frame, width=26)
        self.gvm_user.insert(0, self.config['gvm']['username'])
        self.gvm_user.pack(pady=(0, 4))

        ttk.Label(left_frame, text="GVM Password:").pack(anchor=tk.W)
        self.gvm_pass = ttk.Entry(left_frame, width=26, show="*")
        self.gvm_pass.pack(pady=(0, 8))

        ttk.Label(left_frame, text="Remote Entity / Report Type:", font=("Segoe UI", 9, "bold")).pack(anchor=tk.W)
        self.entity_type_combo = ttk.Combobox(left_frame, values=[
            "📋 Reports (Scan Reports)",
            "⚡ Tasks (Scan Configurations & Schedules)",
            "🔍 Results (Vulnerability Findings)",
            "🛡️ Vulnerabilities (Master NVT Catalog)",
            "📝 Notes (Custom Analyst Notes)",
            "⚖️ Overrides (Severity Adjustment Rules)",
            "🖥️ Hosts (Discovered Asset Inventory)",
            "💻 Operating Systems (OS Inventory)",
            "🔒 TLS Certificates (Discovered SSL/TLS Certificates)"
        ], state="readonly", width=28)
        self.entity_type_combo.current(0)
        self.entity_type_combo.pack(fill=tk.X, pady=(0, 8))

        self.btn_fetch_gvm = ttk.Button(left_frame, text="🔌 Connect & List GVM Items", bootstyle="primary", command=self.fetch_gvm_reports_thread)
        self.btn_fetch_gvm.pack(fill=tk.X, pady=(0, 12))

        ttk.Label(left_frame, text="Export Format:", font=("Segoe UI", 9, "bold")).pack(anchor=tk.W)
        self.format_combo = ttk.Combobox(left_frame, values=list(gvm_connector.FORMATS.keys()), state="readonly")
        self.format_combo.current(0)
        self.format_combo.pack(fill=tk.X, pady=(0, 8))

        self.btn_export_gvm = ttk.Button(left_frame, text="📥 Download & Auto-Load Reports", bootstyle="success", command=self.export_gvm_reports_thread, state=tk.DISABLED)
        self.btn_export_gvm.pack(fill=tk.X, pady=(0, 10))

        ttk.Separator(left_frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=8)

        ttk.Label(left_frame, text="Local Files Controls", font=("Segoe UI", 10, "bold")).pack(anchor=tk.W, pady=(0, 4))
        ttk.Button(left_frame, text="➕ Add Local CSV Files", command=self.load_local_files).pack(fill=tk.X, pady=(0, 4))
        ttk.Button(left_frame, text="🗑️ Remove Selected", command=self.remove_selected_local_file).pack(fill=tk.X, pady=(0, 4))
        ttk.Button(left_frame, text="🔄 Clear Local List", command=self.clear_file_list).pack(fill=tk.X)

        right_frame = ttk.Frame(paned, padding=10)
        paned.add(right_frame, weight=3)

        gvm_hdr = ttk.Frame(right_frame)
        gvm_hdr.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(gvm_hdr, text="Remote OpenVAS Reports (Click rows or checkboxes to select):", font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
        ttk.Button(gvm_hdr, text="🗑️ Delete Selected Remote Report", bootstyle="danger", command=self.delete_gvm_reports_thread).pack(side=tk.RIGHT, padx=4)
        ttk.Button(gvm_hdr, text="☑ Select All", command=self.select_all_gvm).pack(side=tk.RIGHT, padx=2)
        ttk.Button(gvm_hdr, text="☐ Deselect All", command=self.deselect_all_gvm).pack(side=tk.RIGHT, padx=2)

        gvm_tree_frame = ttk.Frame(right_frame)
        gvm_tree_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        gvm_cols = ("Select", "Task Name", "Date", "Report ID")
        self.gvm_tree = ttk.Treeview(gvm_tree_frame, columns=gvm_cols, show="headings", height=8)
        self.gvm_tree.heading("Select", text="☑ Select")
        self.gvm_tree.heading("Task Name", text="Task Name")
        self.gvm_tree.heading("Date", text="Date")
        self.gvm_tree.heading("Report ID", text="Report ID")

        self.gvm_tree.column("Select", width=70, anchor=tk.CENTER)
        self.gvm_tree.column("Task Name", width=240)
        self.gvm_tree.column("Date", width=140)
        self.gvm_tree.column("Report ID", width=220)

        gvm_scroll = ttk.Scrollbar(gvm_tree_frame, orient=tk.VERTICAL, command=self.gvm_tree.yview)
        self.gvm_tree.configure(yscrollcommand=gvm_scroll.set)

        gvm_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.gvm_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.gvm_tree.bind("<Button-1>", self.on_gvm_tree_click)

        local_hdr = ttk.Frame(right_frame)
        local_hdr.pack(fill=tk.X, pady=(5, 5))
        ttk.Label(local_hdr, text="Active Loaded CSV Reports (Checkboxes enabled for multi-scan operations):", font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
        ttk.Button(local_hdr, text="☑ Select All Local", command=self.select_all_local).pack(side=tk.RIGHT, padx=2)
        ttk.Button(local_hdr, text="☐ Deselect All Local", command=self.deselect_all_local).pack(side=tk.RIGHT, padx=2)

        local_tree_frame = ttk.Frame(right_frame)
        local_tree_frame.pack(fill=tk.BOTH, expand=True)

        local_cols = ("Select", "File Path", "File Name")
        self.local_tree = ttk.Treeview(local_tree_frame, columns=local_cols, show="headings", height=6)
        self.local_tree.heading("Select", text="☑ Select")
        self.local_tree.heading("File Path", text="Full Path")
        self.local_tree.heading("File Name", text="File Name")

        self.local_tree.column("Select", width=70, anchor=tk.CENTER)
        self.local_tree.column("File Path", width=380)
        self.local_tree.column("File Name", width=200)

        local_scroll = ttk.Scrollbar(local_tree_frame, orient=tk.VERTICAL, command=self.local_tree.yview)
        self.local_tree.configure(yscrollcommand=local_scroll.set)

        local_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.local_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.local_tree.bind("<Button-1>", self.on_local_tree_click)

    # Checkbox Toggle Handlers
    def on_gvm_tree_click(self, event):
        region = self.gvm_tree.identify("region", event.x, event.y)
        if region == "cell":
            item = self.gvm_tree.identify_row(event.y)
            col = self.gvm_tree.identify_column(event.x)
            if col == "#1" and item:
                vals = list(self.gvm_tree.item(item, "values"))
                vals[0] = "☑" if vals[0] == "☐" else "☐"
                self.gvm_tree.item(item, values=vals)

    def on_local_tree_click(self, event):
        region = self.local_tree.identify("region", event.x, event.y)
        if region == "cell":
            item = self.local_tree.identify_row(event.y)
            col = self.local_tree.identify_column(event.x)
            if col == "#1" and item:
                vals = list(self.local_tree.item(item, "values"))
                vals[0] = "☑" if vals[0] == "☐" else "☐"
                self.local_tree.item(item, values=vals)

    def select_all_gvm(self):
        for item in self.gvm_tree.get_children():
            vals = list(self.gvm_tree.item(item, "values"))
            vals[0] = "☑"
            self.gvm_tree.item(item, values=vals)

    def deselect_all_gvm(self):
        for item in self.gvm_tree.get_children():
            vals = list(self.gvm_tree.item(item, "values"))
            vals[0] = "☐"
            self.gvm_tree.item(item, values=vals)

    def select_all_local(self):
        for item in self.local_tree.get_children():
            vals = list(self.local_tree.item(item, "values"))
            vals[0] = "☑"
            self.local_tree.item(item, values=vals)

    def deselect_all_local(self):
        for item in self.local_tree.get_children():
            vals = list(self.local_tree.item(item, "values"))
            vals[0] = "☐"
            self.local_tree.item(item, values=vals)

    def get_selected_local_file_paths(self):
        checked = []
        all_paths = []
        for item in self.local_tree.get_children():
            vals = self.local_tree.item(item, "values")
            all_paths.append(vals[1])
            if vals[0] == "☑":
                checked.append(vals[1])
        return checked if checked else all_paths

    def load_local_files(self):
        files = filedialog.askopenfilenames(title="Select OpenVAS Report CSVs", filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")])
        for f in files:
            existing = [self.local_tree.item(i, "values")[1] for i in self.local_tree.get_children()]
            if f not in existing:
                self.local_tree.insert("", tk.END, values=("☑", f, os.path.basename(f)))

    def remove_selected_local_file(self):
        for item in list(self.local_tree.get_children()):
            vals = self.local_tree.item(item, "values")
            if vals[0] == "☑":
                self.local_tree.delete(item)

    def clear_file_list(self):
        self.local_tree.delete(*self.local_tree.get_children())

    def fetch_gvm_reports_thread(self):
        if not self.ssh_pass.get() or not self.gvm_pass.get():
            messagebox.showwarning("Missing Credentials", "Please enter both SSH and GVM passwords.")
            return

        self.btn_fetch_gvm.config(state=tk.DISABLED)
        self.start_progress("indeterminate")
        self.log("Establishing SSH Tunnel & Authenticating with GVM...")
        threading.Thread(target=self._fetch_gvm_reports, daemon=True).start()

    def _fetch_gvm_reports(self):
        try:
            reports = gvm_connector.fetch_gvm_reports(
                ssh_host=self.ssh_ip.get(),
                ssh_user=self.ssh_user.get(),
                ssh_pass=self.ssh_pass.get(),
                gvm_user=self.gvm_user.get(),
                gvm_pass=self.gvm_pass.get(),
                entity_type=self.entity_type_combo.get(),
                local_port=self.config['gvm']['local_tunnel_port'],
                remote_socket=self.config['gvm']['socket_path']
            )

            self.gvm_tree.delete(*self.gvm_tree.get_children())
            self.gvm_reports_data.clear()

            for r in reports:
                rid = r['id']
                task = r['task']
                date = r['date']
                self.gvm_reports_data[rid] = r
                self.gvm_tree.insert("", tk.END, iid=rid, values=("☐", task, date, rid))

            # Fetch dynamic report formats (including Certification, SVG, etc.) from GVM server
            try:
                fmt_dict = gvm_connector.fetch_gvm_report_formats(
                    ssh_host=self.ssh_ip.get(),
                    ssh_user=self.ssh_user.get(),
                    ssh_pass=self.ssh_pass.get(),
                    gvm_user=self.gvm_user.get(),
                    gvm_pass=self.gvm_pass.get(),
                    local_port=self.config['gvm']['local_tunnel_port'],
                    remote_socket=self.config['gvm']['socket_path']
                )
                gvm_connector.FORMATS.update(fmt_dict)
                self.format_combo['values'] = list(gvm_connector.FORMATS.keys())
            except Exception:
                pass

            self.log(f"Loaded {len(reports)} remote GVM reports. Target Dir: {REPORTS_DIR}")
            self.btn_export_gvm.config(state=tk.NORMAL)
        except Exception as e:
            self.log(f"Error fetching GVM reports: {e}")
            messagebox.showerror("GVM Connection Error", str(e))
        finally:
            self.stop_progress()
            self.btn_fetch_gvm.config(state=tk.NORMAL)

    def export_gvm_reports_thread(self):
        selected_ids = []
        for item in self.gvm_tree.get_children():
            vals = self.gvm_tree.item(item, "values")
            if vals[0] == "☑":
                selected_ids.append(vals[3])

        if not selected_ids:
            sel = self.gvm_tree.selection()
            if sel: selected_ids = list(sel)

        if not selected_ids:
            messagebox.showinfo("Selection Required", "Please check at least one report checkbox (☑) from the remote table.")
            return

        fmt_name = self.format_combo.get()
        self.btn_export_gvm.config(state=tk.DISABLED)
        self.start_progress("determinate")
        threading.Thread(target=self._export_gvm_reports, args=(selected_ids, fmt_name), daemon=True).start()

    def _export_gvm_reports(self, selected_ids, fmt_name):
        try:
            self.log(f"Connecting to download {len(selected_ids)} report(s) as {fmt_name}...")
            selected_info = [self.gvm_reports_data[rid] for rid in selected_ids if rid in self.gvm_reports_data]
            
            downloaded = gvm_connector.download_gvm_reports(
                ssh_host=self.ssh_ip.get(),
                ssh_user=self.ssh_user.get(),
                ssh_pass=self.ssh_pass.get(),
                gvm_user=self.gvm_user.get(),
                gvm_pass=self.gvm_pass.get(),
                selected_reports=selected_info,
                fmt_name=fmt_name,
                output_dir=REPORTS_DIR,
                local_port=self.config['gvm']['local_tunnel_port'],
                remote_socket=self.config['gvm']['socket_path']
            )

            csv_paths = [p for p in downloaded if p.endswith(".csv")]
            for csv_p in csv_paths:
                existing = [self.local_tree.item(i, "values")[1] for i in self.local_tree.get_children()]
                if csv_p not in existing:
                    self.local_tree.insert("", tk.END, values=("☑", csv_p, os.path.basename(csv_p)))

            self.log(f"✅ Success! Downloaded {len(downloaded)} files to {REPORTS_DIR}")
            messagebox.showinfo("Export Complete", f"Successfully downloaded {len(downloaded)} report file(s) to:\n{REPORTS_DIR}\n\nDownloaded CSVs have been automatically loaded into the active analysis list!")
        except Exception as e:
            self.log("Report download failed.")
            messagebox.showerror("Download Error", str(e))
        finally:
            self.stop_progress()
            self.btn_export_gvm.config(state=tk.NORMAL)

    def delete_gvm_reports_thread(self):
        selected_ids = []
        for item in self.gvm_tree.get_children():
            vals = self.gvm_tree.item(item, "values")
            if vals[0] == "☑":
                selected_ids.append(vals[3])

        if not selected_ids:
            sel = self.gvm_tree.selection()
            if sel: selected_ids = list(sel)

        if not selected_ids:
            messagebox.showinfo("Selection Required", "Please check at least one remote report checkbox (☑) to delete.")
            return

        if not messagebox.askyesno("Confirm Permanent Deletion", f"Are you sure you want to permanently delete {len(selected_ids)} report(s) from OpenVAS?"):
            return

        self.start_progress("indeterminate")
        threading.Thread(target=self._delete_gvm_reports, args=(selected_ids,), daemon=True).start()

    def _delete_gvm_reports(self, selected_ids):
        try:
            self.log(f"Connecting to delete {len(selected_ids)} report(s) from OpenVAS GVM...")
            for rid in selected_ids:
                gvm_connector.delete_gvm_report(
                    ssh_host=self.ssh_ip.get(),
                    ssh_user=self.ssh_user.get(),
                    ssh_pass=self.ssh_pass.get(),
                    gvm_user=self.gvm_user.get(),
                    gvm_pass=self.gvm_pass.get(),
                    report_id=rid,
                    local_port=self.config['gvm']['local_tunnel_port'],
                    remote_socket=self.config['gvm']['socket_path']
                )

            self.log(f"✅ Successfully deleted {len(selected_ids)} remote report(s). Refreshing report list...")
            self._fetch_gvm_reports()
            messagebox.showinfo("Reports Deleted", f"Successfully deleted {len(selected_ids)} report(s) permanently from OpenVAS!")
        except Exception as e:
            self.log("Failed to delete remote reports.")
            messagebox.showerror("Delete Error", str(e))
        finally:
            self.stop_progress()

    # Action Handlers
    def action_send_to_analytics(self):
        paths = self.get_selected_local_file_paths()
        if not paths:
            messagebox.showwarning("No Files", "Please add or select active CSV files first.")
            return
        self.select_view("analytics")
        self.run_risk()

    def action_send_to_trend(self):
        paths = self.get_selected_local_file_paths()
        if not paths:
            messagebox.showwarning("No Files", "Please add or select active CSV files first.")
            return
        self.select_view("trend")
        self.run_trend_analysis()

    def action_send_to_correlation(self):
        paths = self.get_selected_local_file_paths()
        if not paths:
            messagebox.showwarning("No Files", "Please add or select active CSV files first.")
            return
        self.select_view("correlation")
        self.run_correlation()

    def action_send_to_topology(self):
        paths = self.get_selected_local_file_paths()
        if not paths:
            messagebox.showwarning("No Files", "Please add or select active CSV files first.")
            return
        self.select_view("assets")
        self.render_topology()

    def generate_executive_report(self):
        paths = self.get_selected_local_file_paths()
        if not paths:
            messagebox.showwarning("No Active Scans", "Please add or load scan CSV files first.")
            return
        
        save_path = filedialog.asksaveasfilename(
            title="Save Executive Security Summary Report",
            defaultextension=".html",
            initialfile="Executive_Security_Report.html",
            filetypes=[("HTML Reports", "*.html")]
        )
        if save_path:
            try:
                self.start_progress("indeterminate")
                if self.current_risk_df.empty and paths:
                    self.current_risk_df = analyzer.get_risk_score(paths[0], exclude_fp=self.exclude_fp_var.get())
                
                summary_df = analyzer.get_host_risk_summary(self.current_risk_df)
                if self.current_master_df.empty and paths:
                    self.current_master_df = correlation_engine.find_master_targets(paths, fp_filter_fn=lambda df: fp_manager.filter_false_positives(df) if self.exclude_fp_var.get() else df)

                fp_df = fp_manager.get_fp_dataframe()

                rep_file = executive_reporter.generate_executive_html_report(
                    paths, save_path, summary_df, self.current_master_df, fp_df, detailed_df=self.current_risk_df
                )
                messagebox.showinfo("Report Generated", f"Executive HTML report created successfully!\nOpening in default browser:\n{rep_file}")
                webbrowser.open(f"file:///{rep_file}")
            except Exception as e:
                messagebox.showerror("Report Error", f"Failed to generate executive report: {e}")
            finally:
                self.stop_progress()

    # ---------------------------------------------------------
    # TAB 2: ANALYTICS, EXCEL FILTERS, SORTS & DONUT CHART
    # ---------------------------------------------------------
    def setup_analytics_tab(self):
        frame = self.tab_analytics

        # Excel Filter & Sort Control Toolbar
        filter_bar = ttk.LabelFrame(frame, text="📊 Analytics, Threat Intel & Data Controls", padding=8)
        filter_bar.pack(fill=tk.X, padx=15, pady=(10, 5))

        ttk.Button(filter_bar, text="⚡ Run RBVM Scoring", bootstyle="primary", command=self.run_risk).pack(side=tk.LEFT, padx=4)
        ttk.Button(filter_bar, text="🔀 Compare Delta (2 Scans)", bootstyle="secondary", command=self.run_delta).pack(side=tk.LEFT, padx=4)

        self.exclude_fp_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(filter_bar, text="☑ Filter False Positives", variable=self.exclude_fp_var, command=self.filter_analytics_tree).pack(side=tk.LEFT, padx=10)

        # Excel Severity Filter
        ttk.Label(filter_bar, text="Filter Severity:").pack(side=tk.LEFT, padx=(10, 2))
        self.severity_filter_combo = ttk.Combobox(filter_bar, values=["All Severities", "Critical (>=25)", "High (>=10)", "Medium (>=4)", "Low (<4)"], state="readonly", width=16)
        self.severity_filter_combo.current(0)
        self.severity_filter_combo.pack(side=tk.LEFT, padx=2)
        self.severity_filter_combo.bind("<<ComboboxSelected>>", self.filter_analytics_tree)

        # Min Score Filter
        ttk.Label(filter_bar, text="Min Score:").pack(side=tk.LEFT, padx=(10, 2))
        self.min_score_var = tk.StringVar(value="0")
        self.min_score_entry = ttk.Spinbox(filter_bar, from_=0, to=1000, width=5, textvariable=self.min_score_var, command=self.filter_analytics_tree)
        self.min_score_entry.pack(side=tk.LEFT, padx=2)
        self.min_score_entry.bind("<KeyRelease>", self.filter_analytics_tree)

        # Search Bar
        ttk.Label(filter_bar, text="🔍 Search:").pack(side=tk.LEFT, padx=(10, 2))
        self.search_analytics_var = tk.StringVar()
        self.search_analytics_var.trace_add("write", self.filter_analytics_tree)
        search_entry = ttk.Entry(filter_bar, textvariable=self.search_analytics_var, width=16)
        search_entry.pack(side=tk.LEFT, padx=2)

        # Excel Export Button
        ttk.Button(filter_bar, text="📊 Export to Excel (.xlsx)", bootstyle="success", command=self.export_analytics_to_excel).pack(side=tk.RIGHT, padx=4)

        # Statistics Summary Banner
        self.stats_banner_var = tk.StringVar(value="Summary: Total Items: 0 | Critical: 0 | High: 0 | Max Risk Score: 0")
        stats_lbl = ttk.Label(frame, textvariable=self.stats_banner_var, font=("Segoe UI", 9, "bold"), foreground="#2563eb", background="#eff6ff", padding=5, relief=tk.SOLID)
        stats_lbl.pack(fill=tk.X, padx=15, pady=(2, 5))

        # Split View (Left: Treeview with Header Sorting, Right: Donut Severity Breakdown Chart)
        split_frame = ttk.Frame(frame)
        split_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=5)

        tree_frame = ttk.Frame(split_frame)
        tree_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.tree = ttk.Treeview(tree_frame)
        tree_scroll = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.tree.tag_configure('Critical', background='#fee2e2', foreground='#991b1b')
        self.tree.tag_configure('High', background='#ffedd5', foreground='#c2410c')
        self.tree.tag_configure('Medium', background='#fef9c3', foreground='#a16207')
        self.tree.tag_configure('Low', background='#dcfce7', foreground='#15803d')

        self.tree_menu = tk.Menu(self.root, tearoff=0)
        self.tree_menu.add_command(label="🏷️ Tag Selected as False Positive", command=self.context_tag_fp)
        self.tree.bind("<Button-3>", self.show_context_menu)

        self.donut_frame = ttk.Frame(split_frame, width=360)
        self.donut_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=False, padx=(10, 0))

        self.full_analytics_df = pd.DataFrame()

    def render_dataframe(self, df):
        self.full_analytics_df = df.copy()
        self.filter_analytics_tree()

    def filter_analytics_tree(self, *args):
        df = self.full_analytics_df.copy()
        if df.empty:
            self.tree.delete(*self.tree.get_children())
            self.stats_banner_var.set("Summary: Total Items: 0 | Max Risk Score: 0")
            return

        # 1. Apply Severity Dropdown Filter
        sev_filter = self.severity_filter_combo.get()
        if "Risk Score" in df.columns:
            if "Critical" in sev_filter:
                df = df[df['Risk Score'] >= 25]
            elif "High" in sev_filter:
                df = df[df['Risk Score'] >= 10]
            elif "Medium" in sev_filter:
                df = df[df['Risk Score'] >= 4]
            elif "Low" in sev_filter:
                df = df[df['Risk Score'] < 4]

        # 2. Apply Min Score Filter
        try:
            min_score = float(self.min_score_var.get())
            if min_score > 0 and "Risk Score" in df.columns:
                df = df[df['Risk Score'] >= min_score]
        except ValueError:
            pass

        # 3. Apply Search Query Filter
        query = self.search_analytics_var.get().lower().strip()
        if query:
            mask = df.astype(str).apply(lambda row: row.str.lower().str.contains(query).any(), axis=1)
            df = df[mask]

        display_df = df.copy()
        display_df.insert(0, '#', range(1, len(display_df) + 1))

        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = list(display_df.columns)
        self.tree["show"] = "headings"

        # Interactive Sorting Headers
        for col in display_df.columns:
            if col == '#':
                self.tree.heading(col, text='#')
                self.tree.column(col, width=50, anchor=tk.CENTER)
            else:
                self.tree.heading(col, text=col, command=lambda c=col: self.sort_treeview_column(self.tree, c, lambda: df, self.render_dataframe))
                self.tree.column(col, width=140, anchor=tk.W)

        for _, row in display_df.iterrows():
            tag = str(row.get('Severity', '')).capitalize()
            self.tree.insert("", tk.END, values=list(row), tags=(tag,))

        # Update Statistics Summary Banner
        total_cnt = len(df)
        max_score = df['Risk Score'].max() if 'Risk Score' in df.columns and len(df) > 0 else 0
        avg_score = df['Risk Score'].mean() if 'Risk Score' in df.columns and len(df) > 0 else 0
        self.stats_banner_var.set(f"📊 Summary Statistics: Total Records: {total_cnt} | Max Risk Score: {max_score:.1f} | Average Risk Score: {avg_score:.1f}")

    def export_analytics_to_excel(self):
        self.export_dataframe_to_excel(self.full_analytics_df, "Analytics_Risk_Score_Export.xlsx")

    def render_donut_chart(self, paths):
        for widget in self.donut_frame.winfo_children():
            widget.destroy()

        fig, ax = plt.subplots(figsize=(3.8, 3.8), dpi=90)
        fig.patch.set_facecolor('#f8fafc')

        try:
            dfs = []
            if isinstance(paths, str):
                paths = [paths]
            for p in paths:
                raw_df = pd.read_csv(p)
                raw_df = analyzer.normalize_columns(raw_df)
                if self.exclude_fp_var.get():
                    raw_df = fp_manager.filter_false_positives(raw_df)
                dfs.append(raw_df)

            combined_raw = pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()

            if 'Severity' in combined_raw.columns and not combined_raw.empty:
                counts = combined_raw['Severity'].astype(str).str.capitalize().value_counts()
                colors_map = {'Critical': '#ef4444', 'High': '#f97316', 'Medium': '#eab308', 'Low': '#22c55e', 'Log': '#94a3b8'}
                colors = [colors_map.get(k, '#cbd5e1') for k in counts.index]

                ax.pie(counts, labels=counts.index, autopct='%1.1f%%', colors=colors, startangle=140, wedgeprops=dict(width=0.4, edgecolor='white'))
                ax.set_title(f"Severity Breakdown ({len(paths)} Reports)", fontsize=10, fontweight='bold')
            else:
                ax.text(0.5, 0.5, "No Severity column", ha='center', va='center')
                ax.axis('off')
        except Exception:
            ax.text(0.5, 0.5, "Chart Error", ha='center', va='center')
            ax.axis('off')

        canvas = FigureCanvasTkAgg(fig, master=self.donut_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def run_risk(self):
        paths = self.get_selected_local_file_paths()
        if not paths:
            messagebox.showwarning("Select Files", "Please add or check active CSV files first.")
            return
        try:
            self.start_progress("indeterminate")
            if len(paths) == 1:
                self.log(f"Calculating risk score for {os.path.basename(paths[0])}...")
                df = analyzer.get_risk_score(paths[0], exclude_fp=self.exclude_fp_var.get())
            else:
                self.log(f"Combining & calculating risk score across {len(paths)} selected reports...")
                df = analyzer.get_multiple_risk_scores(paths, exclude_fp=self.exclude_fp_var.get())

            self.current_risk_df = df
            self.render_dataframe(df)
            self.render_donut_chart(paths)
            self.log(f"Analyzed risk score across {len(paths)} selected report(s)")
        except Exception as e:
            messagebox.showerror("Analytics Error", f"Failed to calculate risk score: {e}")
        finally:
            self.stop_progress()

    def run_delta(self):
        paths = self.get_selected_local_file_paths()
        if len(paths) < 2:
            messagebox.showwarning("Select 2 Files", "Please check 2 active CSV files to compare.")
            return
        try:
            self.start_progress("indeterminate")
            self.log("Comparing scan deltas...")
            df = analyzer.calculate_delta(paths[0], paths[1], exclude_fp=self.exclude_fp_var.get())
            self.render_dataframe(df)
            self.log("Computed scan delta comparison.")
        except Exception as e:
            messagebox.showerror("Delta Error", f"Failed to compute scan delta: {e}")
        finally:
            self.stop_progress()

    def show_context_menu(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            self.tree_menu.post(event.x_root, event.y_root)

    def context_tag_fp(self):
        sel_item = self.tree.selection()
        if not sel_item:
            return
        values = self.tree.item(sel_item[0], "values")
        cols = self.tree["columns"]
        
        host_idx = cols.index("Hostname") if "Hostname" in cols else 0
        nvt_idx = cols.index("NVT Name") if "NVT Name" in cols else (1 if len(cols) > 1 else 0)

        hostname = str(values[host_idx])
        nvt_name = str(values[nvt_idx])

        reason = simpledialog.askstring("False Positive Tag", f"Enter reason for tagging:\n\nHost: {hostname}\nNVT: {nvt_name}")
        if reason is not None:
            fp_manager.add_fp_tag(hostname, nvt_name, reason)
            messagebox.showinfo("Tagged", f"Vulnerability tagged as False Positive.\n{hostname} - {nvt_name}")
            self.refresh_fp_table()

    # ---------------------------------------------------------
    # TAB 3: FALSE POSITIVES MANAGER & EXCEL EXPORT
    # ---------------------------------------------------------
    def setup_fp_tab(self):
        frame = self.tab_fp

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill=tk.X, padx=15, pady=10)

        ttk.Button(btn_frame, text="➕ Manual Add Tag", command=self.add_manual_fp).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="🗑️ Remove Tag", command=self.remove_fp).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="🔄 Refresh Table", command=self.refresh_fp_table).pack(side=tk.LEFT, padx=4)

        ttk.Button(btn_frame, text="📊 Export FP Audit to Excel", bootstyle="success", command=self.export_fp_to_excel).pack(side=tk.RIGHT, padx=4)

        tree_frame = ttk.Frame(frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)

        self.fp_tree = ttk.Treeview(tree_frame, columns=("Hostname", "NVT Name", "Reason", "Tagged By", "Tagged At"), show="headings")
        fp_scroll = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.fp_tree.yview)
        self.fp_tree.configure(yscrollcommand=fp_scroll.set)

        for col in ("Hostname", "NVT Name", "Reason", "Tagged By", "Tagged At"):
            self.fp_tree.heading(col, text=col, command=lambda c=col: self.sort_treeview_column(self.fp_tree, c, fp_manager.get_fp_dataframe, self.render_fp_dataframe))
            self.fp_tree.column(col, width=150)

        self.fp_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        fp_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.refresh_fp_table()

    def render_fp_dataframe(self, df):
        self.fp_tree.delete(*self.fp_tree.get_children())
        for _, row in df.iterrows():
            self.fp_tree.insert("", tk.END, values=list(row))

    def refresh_fp_table(self):
        df = fp_manager.get_fp_dataframe()
        self.render_fp_dataframe(df)

    def export_fp_to_excel(self):
        df = fp_manager.get_fp_dataframe()
        self.export_dataframe_to_excel(df, "False_Positives_Audit_Log.xlsx")

    def add_manual_fp(self):
        host = simpledialog.askstring("Add Tag", "Enter Hostname or IP:")
        if not host: return
        nvt = simpledialog.askstring("Add Tag", "Enter Vulnerability (NVT) Name:")
        if not nvt: return
        reason = simpledialog.askstring("Add Tag", "Enter Reason:") or "Manual False Positive"

        fp_manager.add_fp_tag(host, nvt, reason)
        self.refresh_fp_table()

    def remove_fp(self):
        sel = self.fp_tree.selection()
        if not sel:
            messagebox.showwarning("Select Tag", "Select a tag to remove.")
            return
        values = self.fp_tree.item(sel[0], "values")
        fp_manager.remove_fp_tag(values[0], values[1])
        self.refresh_fp_table()

    # ---------------------------------------------------------
    # TAB 4: TREND ANALYSIS (HEATMAPS)
    # ---------------------------------------------------------
    def setup_trend_tab(self):
        frame = self.tab_trend

        ctrl_frame = ttk.Frame(frame)
        ctrl_frame.pack(fill=tk.X, padx=15, pady=10)

        ttk.Button(ctrl_frame, text="🔥 Generate Heatmap & Trend (All Active Scans)", command=self.run_trend_analysis).pack(side=tk.LEFT, padx=4)

        ttk.Label(ctrl_frame, text="Select Host for Trajectory:").pack(side=tk.LEFT, padx=(20, 5))
        self.host_combo = ttk.Combobox(ctrl_frame, state="readonly", width=25)
        self.host_combo.pack(side=tk.LEFT, padx=4)
        self.host_combo.bind("<<ComboboxSelected>>", self.on_host_trend_selected)

        ttk.Button(ctrl_frame, text="📊 Export Heatmap Data to Excel", bootstyle="success", command=self.export_trend_to_excel).pack(side=tk.RIGHT, padx=4)

        self.plot_frame = ttk.Frame(frame)
        self.plot_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)

        self.canvas = None
        self.toolbar = None
        self.pivot_df = pd.DataFrame()

    def run_trend_analysis(self):
        files = self.get_selected_local_file_paths()
        if len(files) < 1:
            messagebox.showwarning("No Reports", "Load or check active CSV files first.")
            return

        def fp_filter(df):
            return fp_manager.filter_false_positives(df) if self.exclude_fp_var.get() else df

        self.start_progress("determinate")
        try:
            self.pivot_df, summary_df = trend_engine.aggregate_report_series(
                files,
                fp_filter_fn=fp_filter,
                progress_callback=self.update_progress
            )

            if self.pivot_df.empty:
                messagebox.showerror("Trend Error", "Could not build trend table from loaded files.")
                return

            hosts = sorted(self.pivot_df.index.tolist())
            self.host_combo["values"] = hosts
            if hosts:
                self.host_combo.current(0)

            fig = trend_engine.generate_heatmap_figure(self.pivot_df)
            self.render_plot(fig)
            self.log("Rendered multi-report risk heatmap.")
        finally:
            self.stop_progress()

    def export_trend_to_excel(self):
        if self.pivot_df.empty:
            messagebox.showwarning("No Trend Data", "Generate heatmap trend data first.")
            return
        df_export = self.pivot_df.reset_index()
        self.export_dataframe_to_excel(df_export, "Risk_Score_Trend_Matrix.xlsx")

    def on_host_trend_selected(self, event):
        selected_host = self.host_combo.get()
        if selected_host and not self.pivot_df.empty:
            fig = trend_engine.generate_trend_line_figure(self.pivot_df, selected_host)
            self.render_plot(fig)

    def render_plot(self, fig):
        for widget in self.plot_frame.winfo_children():
            widget.destroy()

        self.canvas = FigureCanvasTkAgg(fig, master=self.plot_frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        self.toolbar = NavigationToolbar2Tk(self.canvas, self.plot_frame)
        self.toolbar.update()
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    # ---------------------------------------------------------
    # TAB 5: MASTER TARGETS CORRELATION & EXCEL FILTERS
    # ---------------------------------------------------------
    def setup_correlation_tab(self):
        frame = self.tab_correlation

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill=tk.X, padx=15, pady=10)

        ttk.Button(btn_frame, text="🌐 Run Master Targets Cross-VLAN Correlation", command=self.run_correlation).pack(side=tk.LEFT, padx=4)

        # Search Bar
        ttk.Label(btn_frame, text="🔍 Search Filter:").pack(side=tk.LEFT, padx=(15, 5))
        self.search_master_var = tk.StringVar()
        self.search_master_var.trace_add("write", self.filter_master_tree)
        search_entry = ttk.Entry(btn_frame, textvariable=self.search_master_var, width=22)
        search_entry.pack(side=tk.LEFT, padx=4)

        # Excel Export Button
        ttk.Button(btn_frame, text="📊 Export Master Targets to Excel", bootstyle="success", command=self.export_master_to_excel).pack(side=tk.RIGHT, padx=4)

        tree_frame = ttk.Frame(frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)

        self.corr_tree = ttk.Treeview(tree_frame)
        corr_scroll = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.corr_tree.yview)
        self.corr_tree.configure(yscrollcommand=corr_scroll.set)

        self.corr_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        corr_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.full_master_df = pd.DataFrame()

    def filter_master_tree(self, *args):
        query = self.search_master_var.get().lower().strip()
        df = self.full_master_df.copy()

        if df.empty:
            self.corr_tree.delete(*self.corr_tree.get_children())
            return

        if query:
            mask = df.astype(str).apply(lambda row: row.str.lower().str.contains(query).any(), axis=1)
            df = df[mask]

        self.render_master_tree(df)

    def render_master_tree(self, df):
        self.corr_tree.delete(*self.corr_tree.get_children())
        self.corr_tree["columns"] = list(df.columns)
        self.corr_tree["show"] = "headings"

        for col in df.columns:
            self.corr_tree.heading(col, text=col, command=lambda c=col: self.sort_treeview_column(self.corr_tree, c, lambda: self.full_master_df, self.render_master_tree))
            self.corr_tree.column(col, width=140, anchor=tk.W)

        for _, row in df.iterrows():
            self.corr_tree.insert("", tk.END, values=list(row))

    def run_correlation(self):
        files = self.get_selected_local_file_paths()
        if not files:
            messagebox.showwarning("No Reports", "Load or check active CSV files first.")
            return

        def fp_filter(df):
            return fp_manager.filter_false_positives(df) if self.exclude_fp_var.get() else df

        self.start_progress("determinate")
        try:
            df = correlation_engine.find_master_targets(
                files,
                fp_filter_fn=fp_filter,
                progress_callback=self.update_progress
            )
            self.current_master_df = df
            self.full_master_df = df
            self.filter_master_tree()
            self.log("Computed Master Targets cross-VLAN correlation.")
        finally:
            self.stop_progress()

    def export_master_to_excel(self):
        self.export_dataframe_to_excel(self.full_master_df, "Master_Targets_Cross_VLAN.xlsx")

    # ---------------------------------------------------------
    # TAB 6: ASSETS & NETWORK TOPOLOGY VISUALIZER (CMDB INVENTORY)
    # ---------------------------------------------------------
    def setup_assets_tab(self):
        frame = self.tab_assets

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill=tk.X, padx=15, pady=8)

        ttk.Button(btn_frame, text="📋 Load CMDB Asset Inventory", bootstyle="primary", command=self.load_and_display_cmdb).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="🗺️ Render Network Topology Map", bootstyle="info", command=self.render_topology).pack(side=tk.LEFT, padx=4)

        ttk.Label(btn_frame, text="🔍 Search Assets:").pack(side=tk.LEFT, padx=(12, 2))
        self.search_cmdb_var = tk.StringVar()
        self.search_cmdb_var.trace_add("write", self.filter_cmdb_tree)
        search_entry = ttk.Entry(btn_frame, textvariable=self.search_cmdb_var, width=18)
        search_entry.pack(side=tk.LEFT, padx=2)

        ttk.Button(btn_frame, text="💾 Export CMDB (CSV/Excel)", bootstyle="success", command=self.export_cmdb).pack(side=tk.RIGHT, padx=4)

        paned = ttk.Panedwindow(frame, orient=tk.VERTICAL)
        paned.pack(fill=tk.BOTH, expand=True, padx=15, pady=5)

        # Top Frame: CMDB Asset Inventory Treeview
        cmdb_frame = ttk.LabelFrame(paned, text="📋 Live CMDB Asset Inventory", padding=6)
        paned.add(cmdb_frame, weight=1)

        tree_scroll_y = ttk.Scrollbar(cmdb_frame, orient=tk.VERTICAL)
        tree_scroll_x = ttk.Scrollbar(cmdb_frame, orient=tk.HORIZONTAL)

        self.cmdb_tree = ttk.Treeview(cmdb_frame, yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set)
        tree_scroll_y.config(command=self.cmdb_tree.yview)
        tree_scroll_x.config(command=self.cmdb_tree.xview)

        tree_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        tree_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.cmdb_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Bottom Frame: Network Topology Canvas
        topo_outer = ttk.LabelFrame(paned, text="🗺️ Interactive Network Topology Map", padding=6)
        paned.add(topo_outer, weight=1)

        self.topo_frame = ttk.Frame(topo_outer)
        self.topo_frame.pack(fill=tk.BOTH, expand=True)

        self.full_cmdb_df = pd.DataFrame()

    def load_and_display_cmdb(self):
        paths = self.get_selected_local_file_paths()
        if not paths:
            messagebox.showwarning("No File Selected", "Please select or check an active CSV file first.")
            return
        path = paths[0]
        try:
            self.start_progress("indeterminate")
            self.log(f"Extracting CMDB Asset Inventory from {os.path.basename(path)}...")
            df = cmdb_exporter.export_assets_from_csv(path)
            self.full_cmdb_df = df
            self.filter_cmdb_tree()
            self.log(f"Loaded {len(df)} CMDB assets from {os.path.basename(path)}")
        except Exception as e:
            messagebox.showerror("CMDB Error", f"Failed to extract CMDB asset inventory: {e}")
        finally:
            self.stop_progress()

    def filter_cmdb_tree(self, *args):
        df = self.full_cmdb_df.copy()
        if df.empty:
            self.cmdb_tree.delete(*self.cmdb_tree.get_children())
            return

        query = self.search_cmdb_var.get().lower().strip()
        if query:
            mask = df.astype(str).apply(lambda row: row.str.lower().str.contains(query).any(), axis=1)
            df = df[mask]

        display_df = df.copy()
        display_df.insert(0, '#', range(1, len(display_df) + 1))

        self.cmdb_tree.delete(*self.cmdb_tree.get_children())
        self.cmdb_tree["columns"] = list(display_df.columns)
        self.cmdb_tree["show"] = "headings"

        for col in display_df.columns:
            if col == '#':
                self.cmdb_tree.heading(col, text='#')
                self.cmdb_tree.column(col, width=50, anchor=tk.CENTER)
            else:
                self.cmdb_tree.heading(col, text=col, command=lambda c=col: self.sort_treeview_column(self.cmdb_tree, c, lambda: self.full_cmdb_df, self.render_cmdb_dataframe))
                self.cmdb_tree.column(col, width=140, anchor=tk.W)

        for _, row in display_df.iterrows():
            self.cmdb_tree.insert("", tk.END, values=list(row))

    def render_cmdb_dataframe(self, df):
        self.full_cmdb_df = df.copy()
        self.filter_cmdb_tree()

    def render_topology(self):
        paths = self.get_selected_local_file_paths()
        if not paths:
            messagebox.showwarning("No File Selected", "Please select or add an active CSV file.")
            return
        path = paths[0]

        def fp_filter(df):
            return fp_manager.filter_false_positives(df) if self.exclude_fp_var.get() else df

        self.start_progress("indeterminate")
        try:
            fig = network_visualizer.build_network_topology_figure(path, fp_filter_fn=fp_filter)

            for widget in self.topo_frame.winfo_children():
                widget.destroy()

            canvas = FigureCanvasTkAgg(fig, master=self.topo_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

            toolbar = NavigationToolbar2Tk(canvas, self.topo_frame)
            toolbar.update()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
            self.log("Rendered NetworkX topology map.")
        finally:
            self.stop_progress()

    def export_cmdb(self):
        paths = self.get_selected_local_file_paths()
        if not paths:
            messagebox.showwarning("No File Selected", "Please select or add an active CSV file to export.")
            return
        path = paths[0]

        save_path = filedialog.asksaveasfilename(
            title="Save CMDB Asset Inventory",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbooks", "*.xlsx"), ("CSV Files", "*.csv")]
        )
        if save_path:
            try:
                self.start_progress("indeterminate")
                df_cmdb = cmdb_exporter.export_assets_from_csv(path)
                if save_path.endswith(".csv"):
                    df_cmdb.to_csv(save_path, index=False)
                else:
                    self.export_dataframe_to_excel(df_cmdb, save_path)
                messagebox.showinfo("Export Successful", f"CMDB Asset inventory saved to:\n{save_path}")
                self.log(f"Exported CMDB asset inventory to {save_path}")
            except Exception as e:
                messagebox.showerror("Export Error", f"Failed to export CMDB inventory: {e}")
            finally:
                self.stop_progress()

    def setup_sla_tab(self):
        """Remediation & SLA Tracking Tab"""
        frame = self.tab_sla
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill=tk.X, padx=15, pady=10)

        ttk.Button(btn_frame, text="⏱️ Analyze SLAs (Requires 2 Scans)", bootstyle="primary", command=self.run_sla_analysis).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="📊 Export SLA Report", bootstyle="success", command=self.export_sla_to_excel).pack(side=tk.RIGHT, padx=4)

        tree_frame = ttk.Frame(frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)
        self.sla_tree = ttk.Treeview(tree_frame)
        sla_scroll = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.sla_tree.yview)
        self.sla_tree.configure(yscrollcommand=sla_scroll.set)
        self.sla_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sla_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.sla_df = pd.DataFrame()

    def run_sla_analysis(self):
        paths = self.get_selected_local_file_paths()
        if len(paths) < 2:
            messagebox.showwarning("Select 2 Files", "Please check exactly 2 active CSV files to compute SLAs (Oldest first recommended).")
            return
        try:
            self.start_progress("indeterminate")
            self.log("Computing Remediation SLAs...")
            # We use delta to find UNRESOLVED vulnerabilities
            df = analyzer.calculate_delta(paths[0], paths[1], exclude_fp=self.exclude_fp_var.get())
            
            # Anything NOT NEW or RESOLVED is basically carried over, but delta function only returns NEW/RESOLVED right now.
            # So let's calculate active vulns in both scans.
            df_old = analyzer.normalize_columns(pd.read_csv(paths[0]))
            df_new = analyzer.normalize_columns(pd.read_csv(paths[1]))
            
            df_old['key'] = df_old['Hostname'].astype(str) + "_" + df_old.get('NVT Name', '').astype(str)
            df_new['key'] = df_new['Hostname'].astype(str) + "_" + df_new.get('NVT Name', '').astype(str)
            
            # Find vulnerabilities that exist in BOTH (i.e. still active)
            active_df = df_new[df_new['key'].isin(df_old['key'])].copy()
            active_df['SLA Status'] = active_df['Severity'].apply(
                lambda s: "Breached (Critical/High)" if s in ['Critical', 'High'] else "Warning"
            )
            
            self.sla_df = active_df[['Hostname', 'Severity', 'NVT Name', 'SLA Status']]
            self.sla_tree.delete(*self.sla_tree.get_children())
            self.sla_tree["columns"] = list(self.sla_df.columns)
            self.sla_tree["show"] = "headings"
            for col in self.sla_df.columns:
                self.sla_tree.heading(col, text=col)
                self.sla_tree.column(col, width=150)
            for _, row in self.sla_df.iterrows():
                self.sla_tree.insert("", tk.END, values=list(row))
            self.log("Computed SLA Breaches.")
        except Exception as e:
            messagebox.showerror("SLA Error", f"Failed to compute SLAs: {e}")
        finally:
            self.stop_progress()

    def export_sla_to_excel(self):
        self.export_dataframe_to_excel(self.sla_df, "SLA_Report.xlsx")

if __name__ == "__main__":
    root = ttk.Window(themename="superhero")
    app = OpenVASApp(root)
    root.mainloop()