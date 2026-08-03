import paramiko
import threading
import socket as socketlib
import time
import base64
import os
from pathlib import Path
import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl.utils import get_column_letter

from gvm.connections._connection import AbstractGvmConnection
from gvm.protocols.gmp import Gmp
from gvm.transforms import EtreeCheckCommandTransform

LOCAL_PORT = 9390
DEFAULT_REMOTE_SOCKET = "/var/lib/docker/volumes/greenbone-community-edition_gvmd_socket_vol/_data/gvmd.sock"

# OpenVAS Official Format UUIDs (Including Certification, Topology SVG, LaTeX, IT-Grundschutz)
FORMATS = {
    "Excel (Parsed from CSV)": "c1645568-627a-11e3-a660-406186ea4fc5",
    "Raw CSV": "c1645568-627a-11e3-a660-406186ea4fc5",
    "PDF Document": "c402cc3e-b531-11e1-9163-406186ea4fc5",
    "HTML Report": "6c248850-1f62-11e1-b082-406186ea4fc5",
    "XML": "a994b278-1f62-11e1-96ac-406186ea4fc5",
    "Plain Text (TXT)": "a3810a62-1f62-11e1-9219-406186ea4fc5",
    "Topology SVG": "910200ca-4e05-11e1-9646-406186ea4fc5",
    "LaTeX Document": "a684c62c-1f62-11e1-a06f-406186ea4fc5",
    "IT-Grundschutz / Certification": "5057e5cc-b825-11e4-9d0e-28924a31e6cd",
    "Verinice ISM": "c15ad349-b0f8-410a-939b-e854816d8438",
    "CPE XML": "8182cb82-1662-11e1-9032-406186ea4fc5"
}

class PlainTCPConnection(AbstractGvmConnection):
    def __init__(self, port=LOCAL_PORT, timeout=60):
        super().__init__(timeout=timeout)
        self.host = "127.0.0.1"
        self.port = port

    def connect(self):
        self._socket = socketlib.socket(socketlib.AF_INET, socketlib.SOCK_STREAM)
        self._socket.settimeout(self._timeout)
        self._socket.connect((self.host, self.port))

class LocalTCPToRemoteUnixSocketTunnel:
    def __init__(self, transport, local_port=LOCAL_PORT, remote_socket_path=DEFAULT_REMOTE_SOCKET, ssh_pass=None):
        self.transport = transport
        self.local_port = local_port
        self.remote_socket_path = remote_socket_path
        self.ssh_pass = ssh_pass
        self._server_sock = None
        self._stop = threading.Event()

    def start(self):
        # 1. Start remote background socat TCP listener bridging to UNIX socket
        try:
            chan = self.transport.open_session()
            if self.ssh_pass:
                cmd = f"echo '{self.ssh_pass}' | sudo -S socat TCP-LISTEN:{self.local_port},bind=127.0.0.1,reuseaddr,fork UNIX-CLIENT:{self.remote_socket_path} >/dev/null 2>&1 &"
            else:
                cmd = f"nohup socat TCP-LISTEN:{self.local_port},bind=127.0.0.1,reuseaddr,fork UNIX-CLIENT:{self.remote_socket_path} >/dev/null 2>&1 &"
            chan.exec_command(cmd)
            time.sleep(0.5)
        except Exception as e:
            print(f"[Tunnel Warning] Remote socat launch response: {e}")

        # 2. Bind local socket and forward binary TCP via SSH direct-tcpip channel (-L)
        self._server_sock = socketlib.socket(socketlib.AF_INET, socketlib.SOCK_STREAM)
        self._server_sock.setsockopt(socketlib.SOL_SOCKET, socketlib.SO_REUSEADDR, 1)
        self._server_sock.bind(("127.0.0.1", self.local_port))
        self._server_sock.listen(5)
        threading.Thread(target=self._accept_loop, daemon=True).start()

    def _accept_loop(self):
        while not self._stop.is_set():
            try:
                client_sock, _ = self._server_sock.accept()
                threading.Thread(target=self._handle_client, args=(client_sock,), daemon=True).start()
            except OSError:
                break

    def _handle_client(self, client_sock):
        try:
            # Use Paramiko direct-tcpip SSH channel for pure binary stream (standard SSH -L)
            channel = self.transport.open_channel(
                "direct-tcpip",
                ("127.0.0.1", self.local_port),
                client_sock.getpeername()
            )
        except Exception:
            client_sock.close()
            return

        def pump(src, dst, half_close):
            try:
                while True:
                    data = src.recv(4096)
                    if not data: break
                    dst.sendall(data)
            except Exception: pass
            finally:
                try: half_close()
                except Exception: pass

        threading.Thread(target=pump, args=(client_sock, channel, channel.shutdown_write), daemon=True).start()
        threading.Thread(target=pump, args=(channel, client_sock, lambda: client_sock.shutdown(socketlib.SHUT_WR)), daemon=True).start()

    def stop(self):
        self._stop.set()
        if self._server_sock:
            try: self._server_sock.close()
            except Exception: pass

def is_port_in_use(port=LOCAL_PORT, host="127.0.0.1"):
    """Checks if a TCP port is currently open and listening locally."""
    try:
        with socketlib.socket(socketlib.AF_INET, socketlib.SOCK_STREAM) as s:
            s.settimeout(1.0)
            return s.connect_ex((host, port)) == 0
    except Exception:
        return False

def start_tunnel(ssh_host, ssh_user, ssh_pass, local_port=LOCAL_PORT, remote_socket=DEFAULT_REMOTE_SOCKET):
    """Establishes Paramiko SSH transport and starts local-to-remote socat tunnel."""
    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    client.connect(hostname=ssh_host, username=ssh_user, password=ssh_pass, port=22)
    tunnel = LocalTCPToRemoteUnixSocketTunnel(client.get_transport(), local_port, remote_socket, ssh_pass=ssh_pass)
    tunnel.start()
    time.sleep(1)
    return client, tunnel

def get_gvm_connection(ssh_host, ssh_user, ssh_pass, local_port=LOCAL_PORT, remote_socket=DEFAULT_REMOTE_SOCKET):
    """
    Smart connection builder:
    If a local tunnel (e.g. background SSH command or existing tunnel) is ALREADY running on local_port,
    it reuses it directly. Otherwise, automatically starts a new Paramiko SSH tunnel.
    """
    if is_port_in_use(local_port):
        print(f"[GVM Connector] Reusing active background SSH tunnel on port {local_port}.")
        return None, None
    else:
        print(f"[GVM Connector] Opening new SSH connection & tunnel to {ssh_host}:{local_port}...")
        return start_tunnel(ssh_host, ssh_user, ssh_pass, local_port, remote_socket)

def fetch_gvm_reports(ssh_host, ssh_user, ssh_pass, gvm_user, gvm_pass, entity_type="🎯 Scan Vulnerability Reports", local_port=LOCAL_PORT, remote_socket=DEFAULT_REMOTE_SOCKET):
    """Connects via active/new tunnel and fetches specified report/entity list from GVM."""
    ssh, tunnel = get_gvm_connection(ssh_host, ssh_user, ssh_pass, local_port, remote_socket)
    reports_list = []
    try:
        gvm_conn = PlainTCPConnection(port=local_port)
        with Gmp(connection=gvm_conn, transform=EtreeCheckCommandTransform()) as gmp:
            gmp.authenticate(gvm_user, gvm_pass)
            
            if "TLS Certificates" in entity_type or "Certificate" in entity_type:
                items = []
                # Try get_tls_certificates, get_certificates, and get_assets(asset_type='tls_certificate')
                for fn_name in ['get_tls_certificates', 'get_certificates']:
                    if hasattr(gmp, fn_name):
                        try:
                            resp = getattr(gmp, fn_name)()
                            found = resp.findall(".//certificate") + resp.findall(".//tls_certificate") + resp.findall(".//asset")
                            if found:
                                items = found
                                break
                        except Exception:
                            pass
                if not items:
                    try:
                        resp = gmp.get_assets(asset_type="tls_certificate")
                        items = resp.findall(".//asset")
                    except Exception:
                        pass

                for c in items:
                    cid = c.get("id") or "N/A"
                    name = c.findtext("name") or c.findtext("subject") or c.findtext("hostname") or f"Cert {cid[:8]}"
                    date = c.findtext("expiration") or c.findtext("modification_time") or c.findtext("creation_time") or "N/A"
                    reports_list.append({"id": cid, "task": f"[Cert] {name}", "date": date})

            elif "Operating Systems" in entity_type:
                items = []
                try:
                    resp = gmp.get_assets(asset_type="os")
                    items = resp.findall(".//asset")
                except Exception:
                    try:
                        resp = gmp.get_assets()
                        items = [a for a in resp.findall(".//asset") if a.get("type") == "os"]
                    except Exception:
                        pass

                for a in items:
                    aid = a.get("id") or "N/A"
                    name = a.findtext("name") or a.findtext("host/name") or f"OS Asset {aid[:8]}"
                    os_name = a.findtext("os/name") or a.findtext("value") or "Discovered OS"
                    date = a.findtext("modification_time") or a.findtext("creation_time") or "N/A"
                    reports_list.append({"id": aid, "task": f"[OS] {name} ({os_name})", "date": date})

            elif "Hosts" in entity_type:
                items = []
                try:
                    resp = gmp.get_assets(asset_type="host")
                    items = resp.findall(".//asset")
                except Exception:
                    try:
                        resp = gmp.get_assets()
                        items = [a for a in resp.findall(".//asset") if a.get("type") in ["host", None]]
                    except Exception:
                        pass

                for a in items:
                    aid = a.get("id") or "N/A"
                    name = a.findtext("name") or a.findtext("host/name") or a.findtext("ip") or f"Host {aid[:8]}"
                    date = a.findtext("modification_time") or a.findtext("creation_time") or "N/A"
                    reports_list.append({"id": aid, "task": f"[Host] {name}", "date": date})

            elif "Results" in entity_type:
                try:
                    resp = gmp.get_results(details=False)
                    items = resp.findall("result")
                except Exception:
                    items = []
                for r in items:
                    rid = r.get("id")
                    name = r.findtext("name") or "Result"
                    host = r.findtext("host") or "N/A"
                    sev = r.findtext("threat") or r.findtext("severity") or "N/A"
                    date = r.findtext("creation_time") or "N/A"
                    reports_list.append({"id": rid, "task": f"[Result] {name} - Host: {host} ({sev})", "date": date})

            elif "Vulnerabilities" in entity_type:
                try:
                    resp = gmp.get_nvts(details=False)
                    items = resp.findall("nvt")
                except Exception:
                    items = []
                for n in items:
                    nid = n.get("oid") or n.get("id")
                    name = n.findtext("name") or "NVT Vulnerability"
                    family = n.findtext("family") or "N/A"
                    reports_list.append({"id": nid, "task": f"[NVT] {name} ({family})", "date": "N/A"})

            elif "Notes" in entity_type:
                try:
                    resp = gmp.get_notes()
                    items = resp.findall("note")
                except Exception:
                    items = []
                for n in items:
                    nid = n.get("id")
                    text = n.findtext("text") or "Note"
                    date = n.findtext("creation_time") or "N/A"
                    reports_list.append({"id": nid, "task": f"[Note] {text[:40]}", "date": date})

            elif "Overrides" in entity_type:
                try:
                    resp = gmp.get_overrides()
                    items = resp.findall("override")
                except Exception:
                    items = []
                for o in items:
                    oid = o.get("id")
                    text = o.findtext("text") or "Override Rule"
                    date = o.findtext("creation_time") or "N/A"
                    reports_list.append({"id": oid, "task": f"[Override] {text[:40]}", "date": date})

            elif "Tasks" in entity_type:
                try:
                    resp = gmp.get_tasks()
                    items = resp.findall("task")
                except Exception:
                    items = []
                for t in items:
                    tid = t.get("id")
                    name = t.findtext("name") or "Task"
                    status = t.findtext("status") or "N/A"
                    date = t.findtext("creation_time") or "N/A"
                    reports_list.append({"id": tid, "task": f"[Task] {name} ({status})", "date": date})

            else:
                # Default: Scan Vulnerability Reports
                resp = gmp.get_reports(details=False)
                reports = resp.findall("report")
                for r in reports:
                    rid = r.get("id")
                    task = r.findtext("task/name") or "Unknown"
                    date = r.findtext("creation_time") or "Unknown"
                    reports_list.append({"id": rid, "task": task, "date": date})

    finally:
        if tunnel: tunnel.stop()
        if ssh: ssh.close()

    return reports_list

def download_gvm_reports(ssh_host, ssh_user, ssh_pass, gvm_user, gvm_pass, selected_reports, fmt_name, output_dir, local_port=LOCAL_PORT, remote_socket=DEFAULT_REMOTE_SOCKET):
    """
    Downloads selected reports in specified format from GVM and saves to output_dir.
    selected_reports: list of dicts [{'id': rid, 'task': task_name}, ...]
    Returns list of downloaded file paths.
    """
    downloaded_paths = []
    fmt_id = FORMATS.get(fmt_name, FORMATS["Raw CSV"])
    out_path = Path(output_dir)
    out_path.mkdir(parents=True, exist_ok=True)

    ssh, tunnel = get_gvm_connection(ssh_host, ssh_user, ssh_pass, local_port, remote_socket)
    try:
        gvm_conn = PlainTCPConnection(port=local_port)
        with Gmp(connection=gvm_conn, transform=EtreeCheckCommandTransform()) as gmp:
            gmp.authenticate(gvm_user, gvm_pass)
            
            for r_info in selected_reports:
                rid = r_info['id']
                task_name = r_info['task']

                response = gmp.get_report(report_id=rid, report_format_id=fmt_id, ignore_pagination=True, details=True)
                report_element = response.find("report")
                
                b64_content = None
                if report_element is not None:
                    if report_element.text and report_element.text.strip():
                        b64_content = report_element.text.strip()
                    if not b64_content:
                        rf = report_element.find("report_format")
                        if rf is not None and rf.tail and rf.tail.strip():
                            b64_content = rf.tail.strip()
                    if not b64_content:
                        texts = [t.strip() for t in report_element.itertext() if t.strip()]
                        if texts:
                            b64_content = max(texts, key=len)

                if not b64_content:
                    continue

                raw_bytes = base64.b64decode(b64_content)
                safe_name = "".join(c for c in task_name if c.isalnum() or c in " -_").strip()
                
                if fmt_name == "Excel (Parsed from CSV)":
                    csv_file = out_path / f"{safe_name}_{rid[:8]}.csv"
                    xlsx_file = out_path / f"{safe_name}_{rid[:8]}.xlsx"
                    csv_file.write_bytes(raw_bytes)
                    downloaded_paths.append(str(csv_file))

                    df = pd.read_csv(csv_file)
                    with pd.ExcelWriter(xlsx_file, engine="openpyxl") as writer:
                        df.to_excel(writer, index=False, sheet_name="Vulnerabilities")
                        ws = writer.sheets["Vulnerabilities"]
                        for cell in ws[1]:
                            cell.font = cell.font.copy(bold=True)
                        for i, col in enumerate(df.columns, start=1):
                            ws.column_dimensions[get_column_letter(i)].width = min(max(df[col].astype(str).map(len).max() if len(df) else 0, len(str(col))) + 2, 60)
                        ws.freeze_panes = "A2"
                    downloaded_paths.append(str(xlsx_file))
                else:
                    ext_map = {
                        "PDF Document": ".pdf",
                        "HTML Report": ".html",
                        "XML": ".xml",
                        "Plain Text (TXT)": ".txt",
                        "Raw CSV": ".csv",
                        "Topology SVG": ".svg",
                        "LaTeX Document": ".tex",
                        "IT-Grundschutz / Certification": ".xml",
                        "Verinice ISM": ".xml",
                        "CPE XML": ".xml"
                    }
                    file_file = out_path / f"{safe_name}_{rid[:8]}{ext_map.get(fmt_name, '.dat')}"
                    file_file.write_bytes(raw_bytes)
                    downloaded_paths.append(str(file_file))

    finally:
        if tunnel: tunnel.stop()
        if ssh: ssh.close()

    return downloaded_paths

def fetch_gvm_report_formats(ssh_host, ssh_user, ssh_pass, gvm_user, gvm_pass, local_port=LOCAL_PORT, remote_socket=DEFAULT_REMOTE_SOCKET):
    """Dynamically queries GVM to discover all installed report formats (including custom certification & SVG formats)."""
    ssh, tunnel = get_gvm_connection(ssh_host, ssh_user, ssh_pass, local_port, remote_socket)
    fmt_dict = dict(FORMATS)
    try:
        gvm_conn = PlainTCPConnection(port=local_port)
        with Gmp(connection=gvm_conn, transform=EtreeCheckCommandTransform()) as gmp:
            gmp.authenticate(gvm_user, gvm_pass)
            resp = gmp.get_report_formats()
            formats = resp.findall("report_format")
            for f in formats:
                fmt_id = f.get("id")
                name = f.findtext("name")
                if fmt_id and name:
                    fmt_dict[name] = fmt_id
    except Exception as e:
        print(f"[GVM Connector] Using default report formats list: {e}")
    finally:
        if tunnel: tunnel.stop()
        if ssh: ssh.close()

    return fmt_dict

def delete_gvm_report(ssh_host, ssh_user, ssh_pass, gvm_user, gvm_pass, report_id, local_port=LOCAL_PORT, remote_socket=DEFAULT_REMOTE_SOCKET):
    """Deletes a report permanently from OpenVAS GVM using GMP."""
    ssh, tunnel = get_gvm_connection(ssh_host, ssh_user, ssh_pass, local_port, remote_socket)
    try:
        gvm_conn = PlainTCPConnection(port=local_port)
        with Gmp(connection=gvm_conn, transform=EtreeCheckCommandTransform()) as gmp:
            gmp.authenticate(gvm_user, gvm_pass)
            gmp.delete_report(report_id=report_id)
    finally:
        if tunnel: tunnel.stop()
        if ssh: ssh.close()